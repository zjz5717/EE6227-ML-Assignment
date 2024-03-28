# -*- coding: UTF-8 -*-
"""
@Project     : ML Assignment 2 
@File        : binary classification tree.py
@Author      : Super_ze
@Date        : 2024/3/28
@Description : 

Copyright © 2024 Super_ze. All Rights Reserved.
"""

import openpyxl
import numpy as np


def get_entropy_impurity(input_data: list) -> float:
    """
    Calculation the entropy impurity.
    :param input_data: Data after divide
    :return: entropy impurity
    """
    group = [0, 0, 0]
    for i in input_data:
        group[i[4] - 1] += 1
    group_sum = group[0] + group[1] + group[2]
    e0 = 0 if group[0] == 0 else -group[0] / group_sum * np.log2(group[0] / group_sum)
    e1 = 0 if group[1] == 0 else -group[1] / group_sum * np.log2(group[1] / group_sum)
    e2 = 0 if group[2] == 0 else -group[2] / group_sum * np.log2(group[2] / group_sum)
    entropy = e0 + e1 + e2
    return entropy


def get_gini_impurity(input_data: list, feature: int, point: float) -> float:
    """
    Calculating the Gini impurity.
    :param list input_data: Data to be segmented
    :param int feature: Features to be segmented
    :param float point: DIVISION POINT
    :return: gini impurity
    """
    left = [0, 0, 0]
    right = [0, 0, 0]

    for i in input_data:
        if i[feature] >= point:
            right[i[4] - 1] += 1
        else:
            left[i[4] - 1] += 1

    left_sum = left[0] + left[1] + left[2]
    right_sum = right[0] + right[1] + right[2]
    left_gini = 0.5 * (1 - (left[0] / left_sum) ** 2 - (left[1] / left_sum) ** 2 - (left[2] / left_sum) ** 2)
    right_gini = 0.5 * (1 - (right[0] / right_sum) ** 2 - (right[1] / right_sum) ** 2 - (right[2] / right_sum) ** 2)
    gini = left_sum / len(input_data) * left_gini + right_sum / len(input_data) * right_gini
    return gini


def get_best_divide_point_of_feature(input_data: list, feature: int) -> float:
    """
    GET OPTIMAL DIVISION POINT.
    :param list input_data: Data to be segmented
    :param int feature: Features to be segmented
    :return: OPTIMAL DIVISION POINT
    """
    feature_list = []
    for i in input_data:
        feature_list.append(i[feature])
    feature_list = list(set(feature_list))
    feature_list.sort()
    min_gini = 999
    point = -1
    for i in range(len(feature_list) - 1):
        feature_divide_point = (feature_list[i] + feature_list[i + 1]) / 2
        g = get_gini_impurity(input_data, feature, feature_divide_point)
        if g < min_gini:
            point = feature_divide_point
            min_gini = g
    if point != -1:
        return point


def divide(input_data: list, feature: int, point: float) -> (list, list):
    """
    Split the input data into two parts.
    :param input_data: Data to be segmented
    :param feature: Features to be segmented
    :param point: OPTIMAL DIVISION POINT
    :return: Two parts of data
    """
    left = []
    right = []
    for i in input_data:
        if i[feature] >= point:
            right.append(i)
        else:
            left.append(i)
    return left, right


def main():
    # read from xlsx files
    train_book = openpyxl.load_workbook("TrainingData.xlsx")
    test_book = openpyxl.load_workbook("TestData.xlsx")
    train_table = train_book["Sheet1"]
    test_table = test_book["Sheet1"]

    # trans training worksheet into list
    train_table_rows = train_table.rows
    total = []
    for i in train_table_rows:
        total.append([i[0].value, i[1].value, i[2].value, i[3].value, i[4].value])

    # check missing value
    for i in total:
        for j in range(4):
            if type(i[j]) is not float and type(i[j]) is not int:
                print(i, " in group %d has a missing value and will be deleted." % i[4])
                total.remove(i)
                break

    # check outlier
    total_np = np.array(total)
    avg = np.sum(total_np, axis=0) / total_np.shape[0]
    std = np.std(total_np, axis=0)
    mid = np.median(total_np, axis=0)
    for i in total:
        for k in range(4):
            if np.abs(i[k] - avg[k]) > 3.5 * std[k]:
                print(i, " in group %d has a outlier and will be replaced by median." % i[4])
                i[k] = mid[k]
                print("the new value is ", i)

    # binary classification tree
    first_point = [get_best_divide_point_of_feature(total, 0), get_best_divide_point_of_feature(total, 1),
                   get_best_divide_point_of_feature(total, 2), get_best_divide_point_of_feature(total, 3)]
    first_point_gini = [get_gini_impurity(total, 0, first_point[0]), get_gini_impurity(total, 1, first_point[1]),
                        get_gini_impurity(total, 2, first_point[2]), get_gini_impurity(total, 3, first_point[3])]

    first_feature = first_point_gini.index(min(first_point_gini))
    first_point = first_point[first_feature]
    print("\n first feature: %d, first point: %f" % (first_feature, first_point))

    second_left, second_right = divide(total, first_feature, first_point)
    print(" entropy impurity for second left group: %f \n entropy impurity for second right group: %f" %
          (get_entropy_impurity(second_left), get_entropy_impurity(second_right)))

    second_right_point = [get_best_divide_point_of_feature(second_right, 0),
                          get_best_divide_point_of_feature(second_right, 1),
                          get_best_divide_point_of_feature(second_right, 2),
                          get_best_divide_point_of_feature(second_right, 3)]
    second_right_point_gini = [get_gini_impurity(second_right, 0, second_right_point[0]),
                               get_gini_impurity(second_right, 1, second_right_point[1]),
                               get_gini_impurity(second_right, 2, second_right_point[2]),
                               get_gini_impurity(second_right, 3, second_right_point[3])]

    second_right_feature = second_right_point_gini.index(min(second_right_point_gini))
    second_right_point = second_right_point[second_right_feature]
    print("\n second right feature: %d, second right point: %f" % (second_right_feature, second_right_point))

    third_left, third_right = divide(second_right, second_right_feature, second_right_point)
    print(" entropy impurity for third left group: %f \n entropy impurity for third right group: %f" %
          (get_entropy_impurity(third_left), get_entropy_impurity(third_right)))

    third_left_point = [get_best_divide_point_of_feature(third_left, 0),
                        get_best_divide_point_of_feature(third_left, 1),
                        get_best_divide_point_of_feature(third_left, 2),
                        get_best_divide_point_of_feature(third_left, 3)]
    third_left_point_gini = [get_gini_impurity(third_left, 0, third_left_point[0]),
                             get_gini_impurity(third_left, 1, third_left_point[1]),
                             get_gini_impurity(third_left, 2, third_left_point[2]),
                             get_gini_impurity(third_left, 3, third_left_point[3])]
    third_left_feature = third_left_point_gini.index(min(third_left_point_gini))
    third_left_point = third_left_point[third_left_feature]
    print("\n third left feature: %d, third left point: %f" % (third_left_feature, third_left_point))

    fourth_left, fourth_right = divide(third_left, third_left_feature, third_left_point)
    print(" entropy impurity for fourth left group: %f \n entropy impurity for fourth right group: %f" %
          (get_entropy_impurity(fourth_left), get_entropy_impurity(fourth_right)))

    # trans test worksheet into list
    test_table_rows = test_table.rows
    class_test = []
    for i in test_table_rows:
        class_test.append([i[0].value, i[1].value, i[2].value, i[3].value])

    # do predict
    for i in class_test:
        if i[first_feature] < first_feature:
            i.append(1)
        else:
            if i[second_right_feature] >= second_right_point:
                i.append(2)
            else:
                i.append(3)

    # output result
    f = openpyxl.Workbook()
    ft = f.active
    for i in range(30):
        for j in range(5):
            ft.cell(row=i+1, column=j+1, value=class_test[i][j])
    f.save("Result.xlsx")


if __name__ == "__main__":
    main()
